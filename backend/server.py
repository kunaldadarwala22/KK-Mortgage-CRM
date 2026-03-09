from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, UploadFile, File, Query
from fastapi.responses import JSONResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt as pyjwt
from bson import ObjectId
import base64
import io
from openai import AsyncOpenAI

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

DEFAULT_ADMIN_EMAIL = "kunalkapadia2212@gmail.com"
DEFAULT_ADMIN_PASSWORD = "Admin2468!!!"
DEFAULT_ADMIN_NAME = "Kunal Kapadia"

SECOND_USER_EMAIL = "kunal.dadarwala22@gmail.com"
SECOND_USER_PASSWORD = "Admin2468!!!"
SECOND_USER_NAME = "Kunal Dadarwala"

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ.get('JWT_SECRET', 'kk-mortgage-solutions-secret-key-2024')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24 * 7

app = FastAPI(title="KK Mortgage Solutions CRM API")

cors_origins_str = os.environ.get('CORS_ORIGINS', 'https://kk-mortgage-crm.pages.dev')
cors_origins = [origin.strip() for origin in cors_origins_str.split(',')]
if 'http://localhost:3000' not in cors_origins:
    cors_origins.append('http://localhost:3000')

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=cors_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)

api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class UserRole:
    ADMIN = "admin"
    ADVISOR = "advisor"
    ADMIN_STAFF = "admin_staff"

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str = UserRole.ADVISOR

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    user_id: str
    email: str
    name: str
    role: str
    picture: Optional[str] = None
    created_at: datetime

class ClientCreate(BaseModel):
    first_name: str
    last_name: str
    dob: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    current_address: Optional[str] = None
    postcode: Optional[str] = None
    security_property_address: Optional[str] = None
    income: Optional[float] = None
    employment_type: Optional[str] = None
    deposit: Optional[float] = None
    credit_issues: bool = False
    credit_issues_notes: Optional[str] = None
    existing_mortgage_balance: Optional[float] = None
    property_price: Optional[float] = None
    loan_amount: Optional[float] = None
    lead_source: Optional[str] = None
    referral_partner_name: Optional[str] = None
    fact_find_complete: bool = False
    vulnerable_customer: bool = False
    advice_type: Optional[str] = None
    gdpr_consent_date: Optional[str] = None
    advisor_id: Optional[str] = None

class ClientResponse(BaseModel):
    client_id: str
    first_name: str
    last_name: str
    dob: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    current_address: Optional[str] = None
    postcode: Optional[str] = None
    security_property_address: Optional[str] = None
    income: Optional[float] = None
    employment_type: Optional[str] = None
    deposit: Optional[float] = None
    credit_issues: bool = False
    credit_issues_notes: Optional[str] = None
    existing_mortgage_balance: Optional[float] = None
    property_price: Optional[float] = None
    loan_amount: Optional[float] = None
    ltv: Optional[float] = None
    lead_source: Optional[str] = None
    referral_partner_name: Optional[str] = None
    fact_find_complete: bool = False
    vulnerable_customer: bool = False
    advice_type: Optional[str] = None
    gdpr_consent_date: Optional[str] = None
    advisor_id: Optional[str] = None
    advisor_name: Optional[str] = None
    created_at: datetime
    updated_at: datetime

class CaseStatus:
    NEW_LEAD = "new_lead"
    FACT_FIND_COMPLETE = "fact_find_complete"
    APPLICATION_SUBMITTED = "application_submitted"
    VALUATION_BOOKED = "valuation_booked"
    OFFER_ISSUED = "offer_issued"
    COMPLETED = "completed"
    LOST_CASE = "lost_case"

class ProductType:
    MORTGAGE = "mortgage"
    INSURANCE = "insurance"

class MortgageType:
    PURCHASE = "purchase"
    REMORTGAGE = "remortgage"
    REMORTGAGE_ADDITIONAL = "remortgage_additional_borrowing"
    PRODUCT_TRANSFER = "product_transfer"

class InsuranceType:
    LIFE = "life_insurance"
    HOME = "home_insurance"
    BUILDINGS = "buildings_insurance"

class CommissionStatus:
    PENDING = "pending"
    SUBMITTED = "submitted_to_lender"
    PAID = "paid"
    CLAWED_BACK = "clawed_back"

class CaseCreate(BaseModel):
    client_id: str
    product_type: str
    mortgage_type: Optional[str] = None
    insurance_type: Optional[str] = None
    status: str = CaseStatus.NEW_LEAD
    term_years: Optional[int] = None
    fixed_rate_period: Optional[int] = None
    interest_rate: Optional[float] = None
    lender_name: Optional[str] = None
    application_reference: Optional[str] = None
    date_application_submitted: Optional[str] = None
    expected_completion_date: Optional[str] = None
    product_start_date: Optional[str] = None
    product_review_date: Optional[str] = None
    product_expiry_date: Optional[str] = None
    loan_amount: Optional[float] = None
    proc_fee_type: Optional[str] = None
    proc_fee_value: Optional[float] = None
    commission_percentage: Optional[float] = None
    gross_commission: Optional[float] = None
    your_commission_share: Optional[float] = None
    proc_fee_total: Optional[float] = None
    commission_status: str = CommissionStatus.PENDING
    advisor_id: Optional[str] = None
    notes: Optional[str] = None

class CaseResponse(BaseModel):
    case_id: str
    client_id: str
    client_name: Optional[str] = None
    product_type: str
    mortgage_type: Optional[str] = None
    insurance_type: Optional[str] = None
    status: str
    term_years: Optional[int] = None
    fixed_rate_period: Optional[int] = None
    interest_rate: Optional[float] = None
    lender_name: Optional[str] = None
    application_reference: Optional[str] = None
    date_application_submitted: Optional[str] = None
    expected_completion_date: Optional[str] = None
    product_start_date: Optional[str] = None
    product_review_date: Optional[str] = None
    product_expiry_date: Optional[str] = None
    loan_amount: Optional[float] = None
    proc_fee_type: Optional[str] = None
    proc_fee_value: Optional[float] = None
    commission_percentage: Optional[float] = None
    gross_commission: Optional[float] = None
    your_commission_share: Optional[float] = None
    proc_fee_total: Optional[float] = None
    commission_status: str
    advisor_id: Optional[str] = None
    advisor_name: Optional[str] = None
    notes: Optional[str] = None
    created_at: datetime
    updated_at: datetime

class TaskCreate(BaseModel):
    title: str
    description: Optional[str] = None
    due_date: str
    client_id: Optional[str] = None
    case_id: Optional[str] = None
    assigned_to: Optional[str] = None
    priority: str = "medium"
    task_type: Optional[str] = None

class TaskResponse(BaseModel):
    task_id: str
    title: str
    description: Optional[str] = None
    due_date: str
    client_id: Optional[str] = None
    client_name: Optional[str] = None
    case_id: Optional[str] = None
    assigned_to: Optional[str] = None
    assigned_to_name: Optional[str] = None
    priority: str
    task_type: Optional[str] = None
    completed: bool = False
    completed_at: Optional[datetime] = None
    created_at: datetime

class DocumentCreate(BaseModel):
    client_id: str
    document_type: str
    file_name: str
    file_data: str
    notes: Optional[str] = None

class DocumentResponse(BaseModel):
    document_id: str
    client_id: str
    document_type: str
    file_name: str
    notes: Optional[str] = None
    uploaded_by: Optional[str] = None
    uploaded_at: datetime

class AuditLogCreate(BaseModel):
    action: str
    entity_type: str
    entity_id: str
    changes: Optional[Dict[str, Any]] = None
    user_id: str

def generate_id(prefix: str = "") -> str:
    return f"{prefix}{uuid.uuid4().hex[:12]}"

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_jwt_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS),
        "iat": datetime.now(timezone.utc)
    }
    return pyjwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_jwt_token(token: str) -> Dict[str, Any]:
    try:
        return pyjwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except pyjwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_current_user(request: Request) -> Dict[str, Any]:
    session_token = request.cookies.get("session_token")
    if session_token:
        session = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
        if session:
            expires_at = session.get("expires_at")
            if isinstance(expires_at, str):
                expires_at = datetime.fromisoformat(expires_at)
            if expires_at.tzinfo is None:
                expires_at = expires_at.replace(tzinfo=timezone.utc)
            if expires_at > datetime.now(timezone.utc):
                user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
                if user:
                    return user

    auth_header = request.headers.get("Authorization")
    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header.split(" ")[1]
        session = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
        if session:
            expires_at = session.get("expires_at")
            if isinstance(expires_at, str):
                expires_at = datetime.fromisoformat(expires_at)
            if expires_at.tzinfo is None:
                expires_at = expires_at.replace(tzinfo=timezone.utc)
            if expires_at > datetime.now(timezone.utc):
                user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
                if user:
                    return user
        try:
            payload = decode_jwt_token(token)
            user = await db.users.find_one({"user_id": payload["user_id"]}, {"_id": 0})
            if user:
                return user
        except:
            pass

    raise HTTPException(status_code=401, detail="Not authenticated")

async def create_audit_log(action: str, entity_type: str, entity_id: str, user_id: str, changes: Dict = None):
    log = {
        "log_id": generate_id("log_"),
        "action": action,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "user_id": user_id,
        "changes": changes,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    await db.audit_logs.insert_one(log)

def calculate_ltv(loan_amount: float, property_price: float) -> float:
    if property_price and property_price > 0:
        return round((loan_amount / property_price) * 100, 2)
    return 0

@api_router.post("/auth/register")
async def register(user: UserCreate):
    existing = await db.users.find_one({"email": user.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    user_id = generate_id("user_")
    hashed = hash_password(user.password)
    user_doc = {
        "user_id": user_id,
        "email": user.email,
        "name": user.name,
        "password": hashed,
        "role": user.role,
        "picture": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    token = create_jwt_token(user_id, user.email, user.role)
    return {"user_id": user_id, "email": user.email, "name": user.name, "role": user.role, "token": token}

@api_router.post("/auth/login")
async def login(user: UserLogin, response: Response):
    db_user = await db.users.find_one({"email": user.email}, {"_id": 0})
    if not db_user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    if not verify_password(user.password, db_user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = create_jwt_token(db_user["user_id"], db_user["email"], db_user["role"])
    session_token = f"session_{uuid.uuid4().hex}"
    session_doc = {
        "session_id": generate_id("sess_"),
        "user_id": db_user["user_id"],
        "session_token": session_token,
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.user_sessions.insert_one(session_doc)
    response.set_cookie(key="session_token", value=session_token, httponly=True, secure=True, samesite="none", path="/", max_age=7 * 24 * 60 * 60)
    return {"user_id": db_user["user_id"], "email": db_user["email"], "name": db_user["name"], "role": db_user["role"], "picture": db_user.get("picture"), "token": token}

@api_router.get("/auth/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    return {"user_id": user["user_id"], "email": user["email"], "name": user["name"], "role": user.get("role", UserRole.ADVISOR), "picture": user.get("picture")}

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    session_token = request.cookies.get("session_token")
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})
    response.delete_cookie(key="session_token", path="/", secure=True, samesite="none")
    return {"message": "Logged out successfully"}

@api_router.get("/users")
async def get_users(request: Request):
    await get_current_user(request)
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    return users

@api_router.get("/users/{user_id}")
async def get_user(user_id: str, request: Request):
    await get_current_user(request)
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, request: Request):
    current_user = await get_current_user(request)
    body = await request.json()
    if current_user["role"] != UserRole.ADMIN and current_user["user_id"] != user_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    update_data = {k: v for k, v in body.items() if k not in ["user_id", "password", "email"]}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.users.update_one({"user_id": user_id}, {"$set": update_data})
    await create_audit_log("update", "user", user_id, current_user["user_id"], update_data)
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password": 0})
    return user

@api_router.post("/clients", status_code=201)
async def create_client(client: ClientCreate, request: Request):
    current_user = await get_current_user(request)
    client_id = generate_id("client_")
    ltv = None
    if client.loan_amount and client.property_price:
        ltv = calculate_ltv(client.loan_amount, client.property_price)
    client_doc = {
        "client_id": client_id,
        **client.model_dump(),
        "ltv": ltv,
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.clients.insert_one(client_doc)
    await create_audit_log("create", "client", client_id, current_user["user_id"])
    result = await db.clients.find_one({"client_id": client_id}, {"_id": 0})
    return result

@api_router.get("/clients/search")
async def search_clients(request: Request, q: Optional[str] = None):
    await get_current_user(request)
    if not q:
        return {"clients": [], "total": 0}
    query = {"$or": [
        {"first_name": {"$regex": q, "$options": "i"}},
        {"last_name": {"$regex": q, "$options": "i"}},
        {"email": {"$regex": q, "$options": "i"}},
        {"phone": {"$regex": q, "$options": "i"}}
    ]}
    clients = await db.clients.find(query, {"_id": 0}).limit(20).to_list(20)
    return {"clients": clients, "total": len(clients)}

@api_router.get("/clients")
async def get_clients(
    request: Request,
    search: Optional[str] = None,
    advisor_id: Optional[str] = None,
    lead_source: Optional[str] = None,
    postcode: Optional[str] = None,
    enrich_cases: bool = True,
    skip: int = 0,
    limit: int = 100
):
    await get_current_user(request)
    query = {}
    if search:
        query["$or"] = [
            {"first_name": {"$regex": search, "$options": "i"}},
            {"last_name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}}
        ]
    if advisor_id:
        query["advisor_id"] = advisor_id
    if lead_source:
        query["lead_source"] = lead_source
    if postcode:
        query["postcode"] = {"$regex": postcode, "$options": "i"}
    clients = await db.clients.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    total = await db.clients.count_documents(query)
    for client in clients:
        if client.get("advisor_id"):
            advisor = await db.users.find_one({"user_id": client["advisor_id"]}, {"_id": 0, "name": 1})
            client["advisor_name"] = advisor["name"] if advisor else None
        if enrich_cases:
            cases = await db.cases.find({"client_id": client["client_id"]}, {"_id": 0}).to_list(50)
            if cases:
                latest = max(cases, key=lambda c: c.get("created_at", ""))
                client["case_status"] = latest.get("status")
                client["commission_status"] = latest.get("commission_status")
                client["expected_completion_date"] = latest.get("expected_completion_date")
                client["case_property_value"] = latest.get("loan_amount")
                client["case_lender"] = latest.get("lender_name")
                client["case_loan_amount"] = latest.get("loan_amount")
                for c in cases:
                    expiry = c.get("product_expiry_date")
                    if expiry:
                        try:
                            exp_date = datetime.strptime(expiry, "%Y-%m-%d")
                            days_until = (exp_date - datetime.now()).days
                            if 0 <= days_until <= 90:
                                client["expiring_soon"] = True
                                break
                        except:
                            pass
            if client.get("loan_amount") and client.get("property_price") and client["property_price"] > 0:
                client["ltv"] = round((client["loan_amount"] / client["property_price"]) * 100, 2)
    return {"clients": clients, "total": total}

@api_router.get("/clients/{client_id}")
async def get_client(client_id: str, request: Request):
    await get_current_user(request)
    client = await db.clients.find_one({"client_id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    if client.get("advisor_id"):
        advisor = await db.users.find_one({"user_id": client["advisor_id"]}, {"_id": 0, "name": 1})
        client["advisor_name"] = advisor["name"] if advisor else None
    return client

@api_router.put("/clients/{client_id}")
async def update_client(client_id: str, request: Request):
    current_user = await get_current_user(request)
    body = await request.json()
    update_data = {k: v for k, v in body.items() if k != "client_id"}
    if "loan_amount" in update_data or "property_price" in update_data:
        existing = await db.clients.find_one({"client_id": client_id}, {"_id": 0})
        loan = update_data.get("loan_amount", existing.get("loan_amount"))
        price = update_data.get("property_price", existing.get("property_price"))
        if loan and price:
            update_data["ltv"] = calculate_ltv(loan, price)
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.clients.update_one({"client_id": client_id}, {"$set": update_data})
    await create_audit_log("update", "client", client_id, current_user["user_id"], update_data)
    client = await db.clients.find_one({"client_id": client_id}, {"_id": 0})
    return client

@api_router.delete("/clients/{client_id}")
async def delete_client(client_id: str, request: Request):
    current_user = await get_current_user(request)
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only admins can delete clients")
    await db.clients.delete_one({"client_id": client_id})
    await db.cases.delete_many({"client_id": client_id})
    await db.documents.delete_many({"client_id": client_id})
    await db.tasks.delete_many({"client_id": client_id})
    await create_audit_log("delete", "client", client_id, current_user["user_id"])
    return {"message": "Client deleted successfully"}

@api_router.post("/cases")
async def create_case(case: CaseCreate, request: Request):
    current_user = await get_current_user(request)
    case_id = generate_id("case_")
    case_doc = {
        "case_id": case_id,
        **case.model_dump(),
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.cases.insert_one(case_doc)
    await create_audit_log("create", "case", case_id, current_user["user_id"])
    await create_case_tasks(case_id, case.status, current_user["user_id"])
    result = await db.cases.find_one({"case_id": case_id}, {"_id": 0})
    return result

@api_router.get("/cases")
async def get_cases(
    request: Request,
    client_id: Optional[str] = None,
    status: Optional[str] = None,
    product_type: Optional[str] = None,
    advisor_id: Optional[str] = None,
    lender_name: Optional[str] = None,
    commission_status: Optional[str] = None,
    skip: int = 0,
    limit: int = 100
):
    await get_current_user(request)
    query = {}
    if client_id:
        query["client_id"] = client_id
    if status:
        query["status"] = status
    if product_type:
        query["product_type"] = product_type
    if advisor_id:
        query["advisor_id"] = advisor_id
    if lender_name:
        query["lender_name"] = {"$regex": lender_name, "$options": "i"}
    if commission_status:
        query["commission_status"] = commission_status
    cases = await db.cases.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    total = await db.cases.count_documents(query)
    for case in cases:
        client = await db.clients.find_one({"client_id": case["client_id"]}, {"_id": 0, "first_name": 1, "last_name": 1})
        case["client_name"] = f"{client['first_name']} {client['last_name']}" if client else None
        if case.get("advisor_id"):
            advisor = await db.users.find_one({"user_id": case["advisor_id"]}, {"_id": 0, "name": 1})
            case["advisor_name"] = advisor["name"] if advisor else None
    return {"cases": cases, "total": total}

@api_router.get("/cases/{case_id}")
async def get_case(case_id: str, request: Request):
    await get_current_user(request)
    case = await db.cases.find_one({"case_id": case_id}, {"_id": 0})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    client = await db.clients.find_one({"client_id": case["client_id"]}, {"_id": 0, "first_name": 1, "last_name": 1})
    case["client_name"] = f"{client['first_name']} {client['last_name']}" if client else None
    if case.get("advisor_id"):
        advisor = await db.users.find_one({"user_id": case["advisor_id"]}, {"_id": 0, "name": 1})
        case["advisor_name"] = advisor["name"] if advisor else None
    return case

@api_router.put("/cases/{case_id}")
async def update_case(case_id: str, request: Request):
    current_user = await get_current_user(request)
    body = await request.json()
    old_case = await db.cases.find_one({"case_id": case_id}, {"_id": 0})
    if not old_case:
        raise HTTPException(status_code=404, detail="Case not found")
    update_data = {k: v for k, v in body.items() if k != "case_id"}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.cases.update_one({"case_id": case_id}, {"$set": update_data})
    await create_audit_log("update", "case", case_id, current_user["user_id"], update_data)
    if "status" in update_data and update_data["status"] != old_case["status"]:
        await create_case_tasks(case_id, update_data["status"], current_user["user_id"])
    case = await db.cases.find_one({"case_id": case_id}, {"_id": 0})
    return case

@api_router.delete("/cases/{case_id}")
async def delete_case(case_id: str, request: Request):
    current_user = await get_current_user(request)
    await db.cases.delete_one({"case_id": case_id})
    await db.tasks.delete_many({"case_id": case_id})
    await create_audit_log("delete", "case", case_id, current_user["user_id"])
    return {"message": "Case deleted successfully"}

async def create_case_tasks(case_id: str, status: str, user_id: str):
    task_templates = {
        CaseStatus.APPLICATION_SUBMITTED: {"title": "Follow up on application", "days": 3},
        CaseStatus.OFFER_ISSUED: {"title": "Review offer with client", "days": 2},
        CaseStatus.COMPLETED: {"title": "Completion follow-up and review", "days": 1}
    }
    if status in task_templates:
        template = task_templates[status]
        task_id = generate_id("task_")
        due_date = (datetime.now(timezone.utc) + timedelta(days=template["days"])).strftime("%Y-%m-%d")
        task_doc = {
            "task_id": task_id,
            "title": template["title"],
            "description": f"Auto-generated task for status: {status}",
            "due_date": due_date,
            "case_id": case_id,
            "assigned_to": user_id,
            "priority": "medium",
            "task_type": "auto",
            "completed": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.tasks.insert_one(task_doc)

@api_router.post("/tasks", status_code=201)
async def create_task(task: TaskCreate, request: Request):
    current_user = await get_current_user(request)
    task_id = generate_id("task_")
    task_doc = {
        "task_id": task_id,
        **task.model_dump(),
        "completed": False,
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.tasks.insert_one(task_doc)
    await create_audit_log("create", "task", task_id, current_user["user_id"])
    result = await db.tasks.find_one({"task_id": task_id}, {"_id": 0})
    return result

@api_router.get("/tasks")
async def get_tasks(
    request: Request,
    assigned_to: Optional[str] = None,
    client_id: Optional[str] = None,
    case_id: Optional[str] = None,
    completed: Optional[bool] = None,
    due_date: Optional[str] = None,
    overdue: Optional[bool] = None,
    skip: int = 0,
    limit: int = 100
):
    await get_current_user(request)
    query = {}
    if assigned_to:
        query["assigned_to"] = assigned_to
    if client_id:
        query["client_id"] = client_id
    if case_id:
        query["case_id"] = case_id
    if completed is not None:
        query["completed"] = completed
    if due_date:
        query["due_date"] = due_date
    if overdue:
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        query["due_date"] = {"$lt": today}
        query["completed"] = False
    tasks = await db.tasks.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    total = await db.tasks.count_documents(query)
    for task in tasks:
        if task.get("client_id"):
            client = await db.clients.find_one({"client_id": task["client_id"]}, {"_id": 0, "first_name": 1, "last_name": 1})
            task["client_name"] = f"{client['first_name']} {client['last_name']}" if client else None
        if task.get("assigned_to"):
            user = await db.users.find_one({"user_id": task["assigned_to"]}, {"_id": 0, "name": 1})
            task["assigned_to_name"] = user["name"] if user else None
    return {"tasks": tasks, "total": total}

@api_router.put("/tasks/{task_id}")
async def update_task(task_id: str, request: Request):
    current_user = await get_current_user(request)
    body = await request.json()
    update_data = {k: v for k, v in body.items() if k != "task_id"}
    if "completed" in update_data and update_data["completed"]:
        update_data["completed_at"] = datetime.now(timezone.utc).isoformat()
    await db.tasks.update_one({"task_id": task_id}, {"$set": update_data})
    await create_audit_log("update", "task", task_id, current_user["user_id"], update_data)
    task = await db.tasks.find_one({"task_id": task_id}, {"_id": 0})
    return task

@api_router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, request: Request):
    current_user = await get_current_user(request)
    await db.tasks.delete_one({"task_id": task_id})
    await create_audit_log("delete", "task", task_id, current_user["user_id"])
    return {"message": "Task deleted successfully"}

@api_router.post("/documents")
async def create_document(doc: DocumentCreate, request: Request):
    current_user = await get_current_user(request)
    doc_id = generate_id("doc_")
    doc_record = {
        "document_id": doc_id,
        "client_id": doc.client_id,
        "document_type": doc.document_type,
        "file_name": doc.file_name,
        "file_data": doc.file_data,
        "notes": doc.notes,
        "uploaded_by": current_user["user_id"],
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    await db.documents.insert_one(doc_record)
    await create_audit_log("create", "document", doc_id, current_user["user_id"])
    return {"document_id": doc_id, "client_id": doc.client_id, "document_type": doc.document_type, "file_name": doc.file_name, "uploaded_at": doc_record["uploaded_at"]}

@api_router.get("/documents")
async def get_documents(request: Request, client_id: Optional[str] = None, document_type: Optional[str] = None):
    await get_current_user(request)
    query = {}
    if client_id:
        query["client_id"] = client_id
    if document_type:
        query["document_type"] = document_type
    documents = await db.documents.find(query, {"_id": 0, "file_data": 0}).to_list(1000)
    return documents

@api_router.get("/documents/{document_id}")
async def get_document(document_id: str, request: Request):
    await get_current_user(request)
    doc = await db.documents.find_one({"document_id": document_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    return doc

@api_router.delete("/documents/{document_id}")
async def delete_document(document_id: str, request: Request):
    current_user = await get_current_user(request)
    await db.documents.delete_one({"document_id": document_id})
    await create_audit_log("delete", "document", document_id, current_user["user_id"])
    return {"message": "Document deleted successfully"}

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(request: Request):
    await get_current_user(request)
    total_clients = await db.clients.count_documents({})
    total_cases = await db.cases.count_documents({})
    pipeline = await db.cases.aggregate([{"$match": {"status": {"$ne": CaseStatus.LOST_CASE}}}, {"$group": {"_id": None, "total_loan": {"$sum": "$loan_amount"}}}]).to_list(1)
    total_pipeline_value = pipeline[0]["total_loan"] if pipeline else 0
    commission_agg = await db.cases.aggregate([{"$match": {"commission_status": CommissionStatus.PAID}}, {"$group": {"_id": None, "total_gross": {"$sum": "$gross_commission"}, "total_proc": {"$sum": "$proc_fee_total"}, "total_share": {"$sum": "$your_commission_share"}}}]).to_list(1)
    commission_stats = commission_agg[0] if commission_agg else {"total_gross": 0, "total_proc": 0, "total_share": 0}
    status_counts = await db.cases.aggregate([{"$group": {"_id": "$status", "count": {"$sum": 1}}}]).to_list(100)
    completed = await db.cases.count_documents({"status": CaseStatus.COMPLETED})
    lost = await db.cases.count_documents({"status": CaseStatus.LOST_CASE})
    conversion_rate = round((completed / total_cases * 100), 2) if total_cases > 0 else 0
    avg_loan = await db.cases.aggregate([{"$match": {"loan_amount": {"$gt": 0}}}, {"$group": {"_id": None, "avg": {"$avg": "$loan_amount"}}}]).to_list(1)
    avg_loan_size = round(avg_loan[0]["avg"], 2) if avg_loan else 0
    ninety_days = (datetime.now(timezone.utc) + timedelta(days=90)).strftime("%Y-%m-%d")
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    expiring_count = await db.cases.count_documents({"product_expiry_date": {"$gte": today, "$lte": ninety_days}, "status": CaseStatus.COMPLETED})
    tasks_today = await db.tasks.count_documents({"due_date": today, "completed": False})
    overdue_tasks = await db.tasks.count_documents({"due_date": {"$lt": today}, "completed": False})
    mortgage_comm = await db.cases.aggregate([{"$match": {"product_type": ProductType.MORTGAGE, "commission_status": CommissionStatus.PAID}}, {"$group": {"_id": None, "total": {"$sum": "$gross_commission"}}}]).to_list(1)
    insurance_comm = await db.cases.aggregate([{"$match": {"product_type": ProductType.INSURANCE, "commission_status": CommissionStatus.PAID}}, {"$group": {"_id": None, "total": {"$sum": "$gross_commission"}}}]).to_list(1)
    return {
        "total_clients": total_clients, "total_cases": total_cases, "total_pipeline_value": total_pipeline_value,
        "total_commission": commission_stats.get("total_gross", 0), "total_proc_fees": commission_stats.get("total_proc", 0),
        "your_commission_share": commission_stats.get("total_share", 0),
        "status_counts": {item["_id"]: item["count"] for item in status_counts},
        "completed_cases": completed, "lost_cases": lost, "conversion_rate": conversion_rate,
        "avg_loan_size": avg_loan_size, "expiring_products": expiring_count,
        "tasks_due_today": tasks_today, "overdue_tasks": overdue_tasks,
        "mortgage_commission": mortgage_comm[0]["total"] if mortgage_comm else 0,
        "insurance_commission": insurance_comm[0]["total"] if insurance_comm else 0
    }

@api_router.get("/dashboard/revenue")
async def get_revenue_analytics(request: Request, start_date: Optional[str] = None, end_date: Optional[str] = None):
    await get_current_user(request)
    monthly_revenue = await db.cases.aggregate([{"$match": {"commission_status": CommissionStatus.PAID}}, {"$addFields": {"month": {"$substr": ["$date_application_submitted", 0, 7]}}}, {"$group": {"_id": "$month", "total": {"$sum": "$gross_commission"}, "proc_fees": {"$sum": "$proc_fee_total"}}}, {"$sort": {"_id": 1}}, {"$limit": 12}]).to_list(12)
    by_lender = await db.cases.aggregate([{"$match": {"commission_status": CommissionStatus.PAID, "lender_name": {"$ne": None}}}, {"$group": {"_id": "$lender_name", "total": {"$sum": "$gross_commission"}, "count": {"$sum": 1}}}, {"$sort": {"total": -1}}, {"$limit": 10}]).to_list(10)
    by_source = await db.cases.aggregate([{"$lookup": {"from": "clients", "localField": "client_id", "foreignField": "client_id", "as": "client"}}, {"$unwind": "$client"}, {"$match": {"commission_status": CommissionStatus.PAID}}, {"$group": {"_id": "$client.lead_source", "total": {"$sum": "$gross_commission"}, "count": {"$sum": 1}}}, {"$sort": {"total": -1}}]).to_list(10)
    by_product = await db.cases.aggregate([{"$match": {"commission_status": CommissionStatus.PAID}}, {"$group": {"_id": "$product_type", "total": {"$sum": "$gross_commission"}, "count": {"$sum": 1}}}]).to_list(10)
    return {"monthly_revenue": monthly_revenue, "by_lender": by_lender, "by_lead_source": by_source, "by_product_type": by_product}

@api_router.get("/dashboard/forecast")
async def get_commission_forecast(request: Request):
    await get_current_user(request)
    today = datetime.now(timezone.utc)
    forecasts = {}
    for days in [30, 60, 90]:
        future_date = (today + timedelta(days=days)).strftime("%Y-%m-%d")
        today_str = today.strftime("%Y-%m-%d")
        result = await db.cases.aggregate([{"$match": {"expected_completion_date": {"$gte": today_str, "$lte": future_date}, "commission_status": {"$in": [CommissionStatus.PENDING, CommissionStatus.SUBMITTED]}}}, {"$group": {"_id": None, "total": {"$sum": "$gross_commission"}, "count": {"$sum": 1}}}]).to_list(1)
        forecasts[f"next_{days}_days"] = {"amount": result[0]["total"] if result else 0, "cases": result[0]["count"] if result else 0}
    return forecasts

@api_router.get("/dashboard/retention")
async def get_retention_data(request: Request):
    await get_current_user(request)
    today = datetime.now(timezone.utc)
    today_str = today.strftime("%Y-%m-%d")
    month_end = (today.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    month_end_str = month_end.strftime("%Y-%m-%d")
    expiring_this_month = await db.cases.find({"product_expiry_date": {"$gte": today_str, "$lte": month_end_str}, "status": CaseStatus.COMPLETED}, {"_id": 0}).to_list(100)
    for case in expiring_this_month:
        client = await db.clients.find_one({"client_id": case["client_id"]}, {"_id": 0, "first_name": 1, "last_name": 1})
        case["client_name"] = f"{client['first_name']} {client['last_name']}" if client else None
    expiring_by_month = await db.cases.aggregate([{"$match": {"product_expiry_date": {"$gte": today_str}, "status": CaseStatus.COMPLETED}}, {"$addFields": {"expiry_month": {"$substr": ["$product_expiry_date", 0, 7]}}}, {"$group": {"_id": "$expiry_month", "count": {"$sum": 1}, "value": {"$sum": "$loan_amount"}}}, {"$sort": {"_id": 1}}, {"$limit": 12}]).to_list(12)
    retention_value = await db.cases.aggregate([{"$match": {"product_expiry_date": {"$gte": today_str}, "status": CaseStatus.COMPLETED}}, {"$group": {"_id": None, "total": {"$sum": "$loan_amount"}}}]).to_list(1)
    return {"expiring_this_month": expiring_this_month, "expiring_by_month": expiring_by_month, "retention_pipeline_value": retention_value[0]["total"] if retention_value else 0}

@api_router.get("/dashboard/lead-analytics")
async def get_lead_analytics(request: Request):
    await get_current_user(request)
    source_stats = await db.cases.aggregate([{"$lookup": {"from": "clients", "localField": "client_id", "foreignField": "client_id", "as": "client"}}, {"$unwind": "$client"}, {"$group": {"_id": "$client.lead_source", "total": {"$sum": 1}, "completed": {"$sum": {"$cond": [{"$eq": ["$status", CaseStatus.COMPLETED]}, 1, 0]}}, "avg_loan": {"$avg": "$loan_amount"}, "total_commission": {"$sum": {"$cond": [{"$eq": ["$commission_status", CommissionStatus.PAID]}, "$gross_commission", 0]}}}}]).to_list(100)
    for stat in source_stats:
        stat["conversion_rate"] = round((stat["completed"] / stat["total"] * 100), 2) if stat["total"] > 0 else 0
    referral_stats = await db.cases.aggregate([{"$lookup": {"from": "clients", "localField": "client_id", "foreignField": "client_id", "as": "client"}}, {"$unwind": "$client"}, {"$match": {"client.lead_source": "referral", "client.referral_partner_name": {"$ne": None}}}, {"$group": {"_id": "$client.referral_partner_name", "count": {"$sum": 1}, "total_commission": {"$sum": {"$cond": [{"$eq": ["$commission_status", CommissionStatus.PAID]}, "$gross_commission", 0]}}}}, {"$sort": {"total_commission": -1}}]).to_list(20)
    return {"by_lead_source": source_stats, "by_referral_partner": referral_stats}

@api_router.get("/audit-logs")
async def get_audit_logs(request: Request, entity_type: Optional[str] = None, entity_id: Optional[str] = None, user_id: Optional[str] = None, skip: int = 0, limit: int = 100):
    current_user = await get_current_user(request)
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only admins can view audit logs")
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    if entity_id:
        query["entity_id"] = entity_id
    if user_id:
        query["user_id"] = user_id
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)
    for log in logs:
        user = await db.users.find_one({"user_id": log["user_id"]}, {"_id": 0, "name": 1})
        log["user_name"] = user["name"] if user else None
    return logs

@api_router.get("/search")
async def global_search(request: Request, q: Optional[str] = None):
    await get_current_user(request)
    if not q or len(q) < 2:
        return {"clients": [], "cases": []}
    client_query = {"$or": [
        {"first_name": {"$regex": q, "$options": "i"}},
        {"last_name": {"$regex": q, "$options": "i"}},
        {"email": {"$regex": q, "$options": "i"}},
        {"phone": {"$regex": q, "$options": "i"}},
    ]}
    clients = await db.clients.find(client_query, {"_id": 0, "client_id": 1, "first_name": 1, "last_name": 1, "email": 1, "phone": 1}).limit(5).to_list(5)
    case_query = {"$or": [
        {"lender_name": {"$regex": q, "$options": "i"}},
        {"case_id": {"$regex": q, "$options": "i"}},
    ]}
    cases = await db.cases.find(case_query, {"_id": 0, "case_id": 1, "client_id": 1, "product_type": 1, "lender_name": 1, "status": 1}).limit(5).to_list(5)
    matching_clients = await db.clients.find(client_query, {"_id": 0, "client_id": 1}).limit(20).to_list(20)
    if matching_clients:
        client_ids = [c["client_id"] for c in matching_clients]
        client_cases = await db.cases.find({"client_id": {"$in": client_ids}}, {"_id": 0, "case_id": 1, "client_id": 1, "product_type": 1, "lender_name": 1, "status": 1}).limit(5).to_list(5)
        for case in client_cases:
            client = await db.clients.find_one({"client_id": case["client_id"]}, {"_id": 0, "first_name": 1, "last_name": 1})
            case["client_name"] = f"{client['first_name']} {client['last_name']}" if client else None
            if not any(c["case_id"] == case["case_id"] for c in cases):
                cases.append(case)
    for case in cases:
        if not case.get("client_name"):
            client = await db.clients.find_one({"client_id": case["client_id"]}, {"_id": 0, "first_name": 1, "last_name": 1})
            case["client_name"] = f"{client['first_name']} {client['last_name']}" if client else None
    return {"clients": clients, "cases": cases[:5]}

@api_router.get("/")
async def root():
    return {"message": "KK Mortgage Solutions CRM API", "version": "1.0"}

@api_router.get("/health")
async def health():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

@api_router.get("/commission/monthly")
async def get_commission_monthly(request: Request, year: Optional[int] = None, start_date: Optional[str] = None, end_date: Optional[str] = None):
    await get_current_user(request)
    match_stage = {}
    if start_date and end_date:
        match_stage["expected_completion_date"] = {"$gte": start_date, "$lte": end_date}
    elif year:
        match_stage["expected_completion_date"] = {"$regex": f"^{year}"}
    pipeline = [
        {"$match": match_stage} if match_stage else {"$match": {}},
        {"$addFields": {"month": {"$cond": {"if": {"$and": [{"$ne": ["$expected_completion_date", None]}, {"$ne": ["$expected_completion_date", ""]}]}, "then": {"$substr": ["$expected_completion_date", 0, 7]}, "else": "unknown"}}}},
        {"$group": {"_id": {"month": "$month", "commission_status": "$commission_status", "product_type": "$product_type"}, "total_commission": {"$sum": {"$ifNull": ["$gross_commission", 0]}}, "total_proc_fees": {"$sum": {"$ifNull": ["$proc_fee_total", 0]}}, "case_count": {"$sum": 1}}},
        {"$sort": {"_id.month": 1}}
    ]
    results = await db.cases.aggregate(pipeline).to_list(500)
    months = {}
    for r in results:
        month = r["_id"]["month"]
        status = r["_id"]["commission_status"] or "pending"
        ptype = r["_id"]["product_type"] or "mortgage"
        if month not in months:
            months[month] = {"month": month, "pending": 0, "submitted": 0, "paid": 0, "clawed_back": 0, "mortgage_commission": 0, "insurance_commission": 0, "proc_fees": 0, "total_commission": 0, "case_count": 0}
        commission = r["total_commission"]
        proc = r["total_proc_fees"]
        status_key = {"pending": "pending", "submitted_to_lender": "submitted", "paid": "paid", "clawed_back": "clawed_back"}.get(status, "pending")
        months[month][status_key] += commission
        if ptype == "mortgage":
            months[month]["mortgage_commission"] += commission
        else:
            months[month]["insurance_commission"] += commission
        months[month]["proc_fees"] += proc
        months[month]["total_commission"] += commission
        months[month]["case_count"] += r["case_count"]
    monthly_data = sorted(months.values(), key=lambda x: x["month"])
    totals = {"total_pending": sum(m["pending"] for m in monthly_data), "total_submitted": sum(m["submitted"] for m in monthly_data), "total_paid": sum(m["paid"] for m in monthly_data), "total_clawed_back": sum(m["clawed_back"] for m in monthly_data), "total_mortgage": sum(m["mortgage_commission"] for m in monthly_data), "total_insurance": sum(m["insurance_commission"] for m in monthly_data), "total_proc_fees": sum(m["proc_fees"] for m in monthly_data), "grand_total": sum(m["total_commission"] for m in monthly_data)}
    return {"monthly": monthly_data, "totals": totals}

@api_router.get("/commission/analytics")
async def get_commission_analytics(request: Request, start_date: Optional[str] = None, end_date: Optional[str] = None, product_filter: Optional[str] = None, commission_status: Optional[str] = None):
    await get_current_user(request)
    match_stage = {}
    if start_date:
        match_stage.setdefault("expected_completion_date", {})["$gte"] = start_date
    if end_date:
        match_stage.setdefault("expected_completion_date", {})["$lte"] = end_date
    if product_filter and product_filter != "all":
        match_stage["product_type"] = product_filter
    if commission_status and commission_status != "all":
        match_stage["commission_status"] = commission_status
    base_match = [{"$match": match_stage}] if match_stage else [{"$match": {}}]
    by_month = await db.cases.aggregate(base_match + [{"$addFields": {"month": {"$cond": {"if": {"$and": [{"$ne": ["$expected_completion_date", None]}, {"$ne": ["$expected_completion_date", ""]}]}, "then": {"$substr": ["$expected_completion_date", 0, 7]}, "else": "unknown"}}}}, {"$group": {"_id": "$month", "total": {"$sum": {"$ifNull": ["$gross_commission", 0]}}, "proc_fees": {"$sum": {"$ifNull": ["$proc_fee_total", 0]}}, "count": {"$sum": 1}}}, {"$sort": {"_id": 1}}, {"$limit": 24}]).to_list(24)
    by_lender = await db.cases.aggregate(base_match + [{"$match": {"lender_name": {"$ne": None}}}, {"$group": {"_id": "$lender_name", "total": {"$sum": {"$ifNull": ["$gross_commission", 0]}}, "proc_fees": {"$sum": {"$ifNull": ["$proc_fee_total", 0]}}, "count": {"$sum": 1}}}, {"$sort": {"total": -1}}, {"$limit": 15}]).to_list(15)
    by_product = await db.cases.aggregate(base_match + [{"$group": {"_id": "$product_type", "total": {"$sum": {"$ifNull": ["$gross_commission", 0]}}, "proc_fees": {"$sum": {"$ifNull": ["$proc_fee_total", 0]}}, "count": {"$sum": 1}}}]).to_list(10)
    by_lead_source = await db.cases.aggregate(base_match + [{"$lookup": {"from": "clients", "localField": "client_id", "foreignField": "client_id", "as": "client"}}, {"$unwind": "$client"}, {"$group": {"_id": "$client.lead_source", "total": {"$sum": {"$ifNull": ["$gross_commission", 0]}}, "proc_fees": {"$sum": {"$ifNull": ["$proc_fee_total", 0]}}, "count": {"$sum": 1}}}, {"$sort": {"total": -1}}]).to_list(20)
    by_advisor = await db.cases.aggregate(base_match + [{"$match": {"advisor_id": {"$ne": None}}}, {"$lookup": {"from": "users", "localField": "advisor_id", "foreignField": "user_id", "as": "advisor"}}, {"$unwind": "$advisor"}, {"$group": {"_id": {"id": "$advisor_id", "name": "$advisor.name"}, "total": {"$sum": {"$ifNull": ["$gross_commission", 0]}}, "proc_fees": {"$sum": {"$ifNull": ["$proc_fee_total", 0]}}, "count": {"$sum": 1}}}, {"$sort": {"total": -1}}]).to_list(20)
    summary = await db.cases.aggregate(base_match + [{"$group": {"_id": None, "total_commission": {"$sum": {"$ifNull": ["$gross_commission", 0]}}, "total_paid": {"$sum": {"$cond": [{"$eq": ["$commission_status", "paid"]}, {"$ifNull": ["$gross_commission", 0]}, 0]}}, "total_pending": {"$sum": {"$cond": [{"$eq": ["$commission_status", "pending"]}, {"$ifNull": ["$gross_commission", 0]}, 0]}}, "total_clawbacks": {"$sum": {"$cond": [{"$eq": ["$commission_status", "clawed_back"]}, {"$ifNull": ["$gross_commission", 0]}, 0]}}, "total_proc_fees": {"$sum": {"$ifNull": ["$proc_fee_total", 0]}}, "case_count": {"$sum": 1}, "avg_commission": {"$avg": {"$ifNull": ["$gross_commission", 0]}}}}]).to_list(1)
    summary_data = summary[0] if summary else {"total_commission": 0, "total_paid": 0, "total_pending": 0, "total_clawbacks": 0, "total_proc_fees": 0, "case_count": 0, "avg_commission": 0}
    if "_id" in summary_data:
        del summary_data["_id"]
    return {"by_month": by_month, "by_lender": by_lender, "by_product": by_product, "by_lead_source": by_lead_source, "by_advisor": [{"name": r["_id"]["name"], "advisor_id": r["_id"]["id"], "total": r["total"], "proc_fees": r["proc_fees"], "count": r["count"]} for r in by_advisor], "summary": summary_data}

@api_router.get("/analytics/mortgage-types")
async def get_mortgage_type_analytics(request: Request):
    await get_current_user(request)
    pipeline = [{"$match": {"product_type": "mortgage", "mortgage_type": {"$ne": None}}}, {"$group": {"_id": "$mortgage_type", "case_count": {"$sum": 1}, "total_commission": {"$sum": {"$cond": [{"$eq": ["$commission_status", "paid"]}, {"$ifNull": ["$gross_commission", 0]}, 0]}}, "total_loan": {"$sum": {"$ifNull": ["$loan_amount", 0]}}, "avg_loan": {"$avg": {"$ifNull": ["$loan_amount", 0]}}}}]
    results = await db.cases.aggregate(pipeline).to_list(10)
    total_cases = sum(r["case_count"] for r in results) or 1
    types = [{"mortgage_type": r["_id"], "case_count": r["case_count"], "percentage": round((r["case_count"] / total_cases) * 100, 1), "total_commission": r["total_commission"], "total_loan": r["total_loan"], "avg_loan": round(r["avg_loan"], 2) if r["avg_loan"] else 0} for r in results]
    return {"types": types, "total_cases": total_cases}

@api_router.get("/reports/cases-completed")
async def get_cases_completed_report(request: Request, start_date: str = Query(...), end_date: str = Query(...)):
    await get_current_user(request)
    cases = await db.cases.find({"status": CaseStatus.COMPLETED, "expected_completion_date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}).to_list(1000)
    for case in cases:
        client = await db.clients.find_one({"client_id": case["client_id"]}, {"_id": 0, "first_name": 1, "last_name": 1, "property_price": 1})
        case["client_name"] = f"{client['first_name']} {client['last_name']}" if client else "Unknown"
        case["property_value"] = client.get("property_price") if client else None
        if case.get("loan_amount") and case.get("property_value") and case["property_value"] > 0:
            case["ltv"] = round((case["loan_amount"] / case["property_value"]) * 100, 2)
        else:
            case["ltv"] = None
    total_loan = sum(c.get("loan_amount", 0) or 0 for c in cases)
    total_commission = sum(c.get("gross_commission", 0) or 0 for c in cases)
    return {"cases": cases, "summary": {"total_cases": len(cases), "total_loan_value": total_loan, "total_commission": total_commission}}

@api_router.get("/reports/commission-paid")
async def get_commission_paid_report(request: Request, start_date: str = Query(...), end_date: str = Query(...)):
    await get_current_user(request)
    cases = await db.cases.find({"commission_status": CommissionStatus.PAID, "expected_completion_date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}).to_list(1000)
    for case in cases:
        client = await db.clients.find_one({"client_id": case["client_id"]}, {"_id": 0, "first_name": 1, "last_name": 1})
        case["client_name"] = f"{client['first_name']} {client['last_name']}" if client else "Unknown"
    total_commission = sum(c.get("gross_commission", 0) or 0 for c in cases)
    total_proc_fees = sum(c.get("proc_fee_total", 0) or 0 for c in cases)
    return {"cases": cases, "summary": {"total_cases": len(cases), "total_commission_paid": total_commission, "total_proc_fees": total_proc_fees, "total_combined_revenue": total_commission + total_proc_fees}}

@api_router.get("/reports/export")
async def export_report(request: Request, report_type: str = Query(...), start_date: str = Query(...), end_date: str = Query(...), format: str = Query(default="xlsx")):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    import csv
    current_user = await get_current_user(request)
    if report_type == "cases_completed":
        data = await get_cases_completed_report(request, start_date, end_date)
        cases = data["cases"]
        summary = data["summary"]
        title = "Cases Completed Report"
        headers = ["Client Name", "Loan Amount", "Property Value", "LTV %", "Lender", "Product Type", "Mortgage Type", "Completion Date", "Commission"]
        def row_fn(c):
            return [c.get("client_name", ""), c.get("loan_amount", ""), c.get("property_value", ""), f"{c.get('ltv', '')}%" if c.get('ltv') else "", c.get("lender_name", ""), (c.get("product_type", "") or "").replace("_", " ").title(), (c.get("mortgage_type", "") or "").replace("_", " ").title(), c.get("expected_completion_date", ""), c.get("gross_commission", "")]
    else:
        data = await get_commission_paid_report(request, start_date, end_date)
        cases = data["cases"]
        summary = data["summary"]
        title = "Commission Paid Report"
        headers = ["Client Name", "Loan Amount", "Lender", "Product Type", "Commission Amount", "Proc Fee", "Payment Date"]
        def row_fn(c):
            return [c.get("client_name", ""), c.get("loan_amount", ""), c.get("lender_name", ""), (c.get("product_type", "") or "").replace("_", " ").title(), c.get("gross_commission", ""), c.get("proc_fee_total", ""), c.get("expected_completion_date", "")]
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for c in cases:
            writer.writerow(row_fn(c))
        output.seek(0)
        return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={title.replace(' ', '_')}_{start_date}_to_{end_date}.csv"})
    wb = Workbook()
    ws = wb.active
    ws.title = title
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    title_cell = ws['A1']
    title_cell.value = f"{title} ({start_date} to {end_date})"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for idx, c in enumerate(cases):
        for col, val in enumerate(row_fn(c), 1):
            cell = ws.cell(row=3 + idx, column=col, value=val)
            if isinstance(val, (int, float)) and val:
                cell.number_format = '£#,##0'
    sum_row = 3 + len(cases) + 1
    ws.cell(row=sum_row, column=1, value="TOTALS").font = Font(bold=True)
    for key, val in summary.items():
        if key != "total_cases":
            ws.cell(row=sum_row + list(summary.keys()).index(key), column=1, value=key.replace("_", " ").title()).font = Font(bold=True)
            ws.cell(row=sum_row + list(summary.keys()).index(key), column=2, value=val).number_format = '£#,##0' if isinstance(val, (int, float)) else ''
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={title.replace(' ', '_')}_{start_date}_to_{end_date}.xlsx"})

@api_router.get("/export/excel")
async def export_all_data(request: Request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    current_user = await get_current_user(request)
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws_clients = wb.active
    ws_clients.title = "Clients"
    client_headers = ["Client ID", "First Name", "Last Name", "Email", "Phone", "DOB", "Address", "Postcode", "Income", "Employment Type", "Deposit", "Property Price", "Loan Amount", "LTV %", "Credit Issues", "Lead Source", "Referral Partner", "Fact Find Complete", "Vulnerable Customer", "Advice Type", "GDPR Consent Date", "Created At"]
    ws_clients.append(client_headers)
    for col, header in enumerate(client_headers, 1):
        cell = ws_clients.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    clients = await db.clients.find({}, {"_id": 0}).to_list(10000)
    for client in clients:
        ws_clients.append([client.get("client_id", ""), client.get("first_name", ""), client.get("last_name", ""), client.get("email", ""), client.get("phone", ""), client.get("dob", ""), client.get("current_address", ""), client.get("postcode", ""), client.get("income", ""), client.get("employment_type", ""), client.get("deposit", ""), client.get("property_price", ""), client.get("loan_amount", ""), client.get("ltv", ""), "Yes" if client.get("credit_issues") else "No", client.get("lead_source", ""), client.get("referral_partner_name", ""), "Yes" if client.get("fact_find_complete") else "No", "Yes" if client.get("vulnerable_customer") else "No", client.get("advice_type", ""), client.get("gdpr_consent_date", ""), client.get("created_at", "")])
    for col in ws_clients.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws_clients.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    ws_cases = wb.create_sheet("Cases")
    case_headers = ["Case ID", "Client ID", "Product Type", "Mortgage Type", "Insurance Type", "Status", "Lender Name", "Loan Amount", "Term (Years)", "Interest Rate", "Application Reference", "Application Date", "Expected Completion", "Product Start Date", "Product Review Date", "Product Expiry Date", "Proc Fee Type", "Proc Fee Value", "Commission %", "Gross Commission", "Your Share", "Proc Fee Total", "Commission Status", "Created At"]
    ws_cases.append(case_headers)
    for col, header in enumerate(case_headers, 1):
        cell = ws_cases.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    cases = await db.cases.find({}, {"_id": 0}).to_list(10000)
    for case in cases:
        ws_cases.append([case.get("case_id", ""), case.get("client_id", ""), case.get("product_type", ""), case.get("mortgage_type", ""), case.get("insurance_type", ""), case.get("status", ""), case.get("lender_name", ""), case.get("loan_amount", ""), case.get("term_years", ""), case.get("interest_rate", ""), case.get("application_reference", ""), case.get("date_application_submitted", ""), case.get("expected_completion_date", ""), case.get("product_start_date", ""), case.get("product_review_date", ""), case.get("product_expiry_date", ""), case.get("proc_fee_type", ""), case.get("proc_fee_value", ""), case.get("commission_percentage", ""), case.get("gross_commission", ""), case.get("your_commission_share", ""), case.get("proc_fee_total", ""), case.get("commission_status", ""), case.get("created_at", "")])
    for col in ws_cases.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws_cases.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    await create_audit_log("export", "all_data", "excel", current_user["user_id"])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"KK_Mortgage_CRM_Export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

@api_router.post("/extract/client")
async def extract_client_from_screenshots(request: Request, files: List[UploadFile] = File(...)):
    await get_current_user(request)
    openai_api_key = os.environ.get("OPENAI_API_KEY")
    if not openai_api_key:
        raise HTTPException(status_code=500, detail="OpenAI API key not configured")
    openai_client = AsyncOpenAI(api_key=openai_api_key)
    image_contents = []
    for file in files:
        data = await file.read()
        b64 = base64.b64encode(data).decode("utf-8")
        mime = file.content_type or "image/png"
        image_contents.append({"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}", "detail": "high"}})
    prompt = """You are a mortgage CRM data extraction assistant. Extract ALL applicant information from these screenshots and return ONLY a valid JSON object with no extra text.

IMPORTANT: If the screenshot contains multiple applicants (e.g. Applicant 1 and Applicant 2, or joint application), extract ALL of them into the applicants array. The first entry is the primary applicant.

Return this structure:
{
  "applicants": [
    {
      "first_name": "",
      "last_name": "",
      "dob": "YYYY-MM-DD format if found, else empty string",
      "phone": "",
      "email": "",
      "current_address": "",
      "postcode": "",
      "income": null,
      "employment_type": "one of: employed, self_employed, contractor, retired, unemployed or empty string",
      "credit_issues": false,
      "credit_issues_notes": "",
      "existing_mortgage_balance": null,
      "lead_source": "",
      "advice_type": ""
    }
  ]
}

If there is only one applicant, still return the applicants array with one entry.
Return null for missing numeric fields, empty string for missing text fields. Return ONLY the JSON object."""
    messages = [{"role": "user", "content": [{"type": "text", "text": prompt}] + image_contents}]
    try:
        response = await openai_client.chat.completions.create(model="gpt-4o", messages=messages, max_tokens=2000, temperature=0)
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        import json
        extracted = json.loads(raw.strip())
        return {"success": True, "data": extracted}
    except Exception as e:
        logger.error(f"OpenAI extraction error: {e}")
        raise HTTPException(status_code=500, detail=f"Extraction failed: {str(e)}")

@api_router.post("/extract/case")
async def extract_case_from_screenshots(request: Request, files: List[UploadFile] = File(...)):
    await get_current_user(request)
    openai_api_key = os.environ.get("OPENAI_API_KEY")
    if not openai_api_key:
        raise HTTPException(status_code=500, detail="OpenAI API key not configured")
    openai_client = AsyncOpenAI(api_key=openai_api_key)
    image_contents = []
    for file in files:
        data = await file.read()
        b64 = base64.b64encode(data).decode("utf-8")
        mime = file.content_type or "image/png"
        image_contents.append({"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}", "detail": "high"}})
    prompt = """You are a mortgage CRM data extraction assistant. Extract mortgage/insurance case information from these screenshots and return ONLY a valid JSON object with no extra text.

Extract ALL of these fields if present:
{
  "product_type": "mortgage or insurance",
  "mortgage_type": "one of: purchase, remortgage, product_transfer, remortgage_additional_borrowing or empty string",
  "lender_name": "",
  "loan_amount": null,
  "property_value": null,
  "deposit": null,
  "deposit_source": "",
  "ltv": null,
  "term_years": null,
  "initial_product_term": null,  // look for "Product Term", "Initial Term", "Fixed Term" or similar
  "interest_rate": null,
  "interest_rate_type": "one of: fixed, variable, discounted, tracker, capped or empty string",
  "repayment_type": "one of: repayment, interest_only or empty string",
  "property_type": "one of: residential, buy_to_let or empty string",
  "case_reference": "",
  "security_address": "",
  "security_postcode": "",
  "date_application_submitted": "YYYY-MM-DD or empty string",
  "expected_completion_date": "YYYY-MM-DD or empty string",
  "product_start_date": "YYYY-MM-DD or empty string",
  "product_expiry_date": "YYYY-MM-DD or empty string",
  "proc_fee_total": null,
  "commission_percentage": null,
  "client_fee": null,
  "insurance_type": "one of: life_insurance, buildings_insurance, home_insurance or empty string",
  "insurance_provider": "",
  "insurance_cover_type": "one of: level_term, decreasing_term, increasing_term, whole_of_life or empty string",
  "monthly_premium": null,
  "sum_assured": null,
  "notes": ""
}

Return null for missing numeric fields, empty string for missing text fields. Return ONLY the JSON object."""
    messages = [{"role": "user", "content": [{"type": "text", "text": prompt}] + image_contents}]
    try:
        response = await openai_client.chat.completions.create(model="gpt-4o", messages=messages, max_tokens=2000, temperature=0)
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        import json
        extracted = json.loads(raw.strip())
        return {"success": True, "data": extracted}
    except Exception as e:
        logger.error(f"OpenAI extraction error: {e}")
        raise HTTPException(status_code=500, detail=f"Extraction failed: {str(e)}")

@api_router.get("/notifications")
async def get_notifications(request: Request):
    await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    in_7_days = (datetime.now(timezone.utc) + timedelta(days=7)).strftime("%Y-%m-%d")
    in_90_days = (datetime.now(timezone.utc) + timedelta(days=90)).strftime("%Y-%m-%d")
    notifications = []
    overdue_tasks = await db.tasks.find({"due_date": {"$lt": today}, "completed": False}, {"_id": 0, "title": 1, "due_date": 1, "task_id": 1}).to_list(10)
    for t in overdue_tasks:
        notifications.append({"type": "overdue_task", "message": f"Overdue task: {t['title']} (due {t['due_date']})", "entity_id": t["task_id"], "severity": "high"})
    upcoming_tasks = await db.tasks.find({"due_date": {"$gte": today, "$lte": in_7_days}, "completed": False}, {"_id": 0, "title": 1, "due_date": 1, "task_id": 1}).to_list(10)
    for t in upcoming_tasks:
        notifications.append({"type": "upcoming_task", "message": f"Task due soon: {t['title']} (due {t['due_date']})", "entity_id": t["task_id"], "severity": "medium"})
    expiring = await db.cases.find({"product_expiry_date": {"$gte": today, "$lte": in_90_days}, "status": "completed"}, {"_id": 0, "case_id": 1, "client_id": 1, "product_expiry_date": 1, "lender_name": 1}).to_list(10)
    for c in expiring:
        client = await db.clients.find_one({"client_id": c["client_id"]}, {"_id": 0, "first_name": 1, "last_name": 1})
        name = f"{client['first_name']} {client['last_name']}" if client else "Unknown"
        notifications.append({"type": "expiring_product", "message": f"Product expiring: {name} with {c.get('lender_name', 'unknown lender')} on {c['product_expiry_date']}", "entity_id": c["case_id"], "severity": "medium"})
    return {"notifications": notifications, "count": len(notifications)}

app.include_router(api_router)

@app.on_event("startup")
async def seed_default_user():
    for email, name, password in [
        (DEFAULT_ADMIN_EMAIL, DEFAULT_ADMIN_NAME, DEFAULT_ADMIN_PASSWORD),
        (SECOND_USER_EMAIL, SECOND_USER_NAME, SECOND_USER_PASSWORD),
    ]:
        existing = await db.users.find_one({"email": email})
        if not existing:
            user_doc = {"user_id": generate_id("user_"), "email": email, "name": name, "password": hash_password(password), "role": UserRole.ADVISOR, "picture": None, "created_at": datetime.now(timezone.utc).isoformat()}
            await db.users.insert_one(user_doc)
            logger.info(f"Default user created: {email}")
        elif not existing.get("password"):
            await db.users.update_one({"email": email}, {"$set": {"password": hash_password(password)}})
            logger.info(f"Fixed missing password for: {email}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
